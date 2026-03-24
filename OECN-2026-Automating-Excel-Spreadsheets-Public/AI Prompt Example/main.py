import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os

def update_tax_balance_sheet():
    # --- STEP 1: LOAD THE PAYROLL DATA ---
    # We use pandas to read the CSV and find the 'FED TAX' row
    input_file = "Payroll Item Summary.csv"
    output_file = "output.xlsx"
    
    try:
        df = pd.read_csv(input_file)
        
        # Filter the dataframe where 'Description' is 'FED TAX'
        # We then extract the 'Employee Share' value from that row
        fed_tax_row = df[df['Description'] == 'FED TAX']
        
        if fed_tax_row.empty:
            print("Error: 'FED TAX' not found in the Description column.")
            return

        # Get the dollar amount (assuming it's the first match)
        tax_amount = fed_tax_row['Employee Share'].values[0]
        
    except Exception as e:
        print(f"Failed to read input file: {e}")
        return

    # --- STEP 2: PREPARE THE OUTPUT FILE ---
    # If the file doesn't exist, we create it with headers July - June
    months = ["July", "August", "September", "October", "November", "December", 
              "January", "February", "March", "April", "May", "June"]

    if not os.path.exists(output_file):
        new_wb = openpyxl.Workbook()
        ws = new_wb.active
        # Start headers in Column B (index 2)
        for i, month in enumerate(months, start=2):
            ws.cell(row=1, column=i).value = month
        new_wb.save(output_file)

    # --- STEP 3: FIND THE NEXT AVAILABLE MONTH ---
    # Load the existing workbook to find where to write the data
    wb = load_workbook(output_file)
    ws = wb.active

    # Loop through columns B through M (2 to 13) to find the first empty cell in Row 2
    data_written = False
    for col in range(2, 14):
        if ws.cell(row=2, column=col).value is None:
            ws.cell(row=2, column=col).value = tax_amount
            current_month = ws.cell(row=1, column=col).value
            print(f"Successfully added {tax_amount} to {current_month}.")
            data_written = True
            break

    if not data_written:
        print("Sheet is full! All months from July to June already have data.")

    # --- STEP 4: SAVE THE RESULTS ---
    wb.save(output_file)

if __name__ == "__main__":
    update_tax_balance_sheet()