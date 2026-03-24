#steps to activate virtual environment:
#1. Open bash command terminal
#2. cd into venv/Scripts
#3. Enter command: source activate or activate.ps1 if using powershell terminal
#cd back into project's main folder

#to run the actual script  enter this command in bash terminal: python main.py


#to install modules to project, make sure venv is activated. CD to main project's directory and run this command as an example: python -m pip install module name
#Remember to use the -m otherwise the module will not install into the virtual environment but the global python environment

import os

#dataframe and excel libraries
import pandas as pd
from pandas import DataFrame
import openpyxl
from openpyxl import Workbook, worksheet
from openpyxl.utils.dataframe import dataframe_to_rows

#custom libraries
from utils.data_processing import format_input_file
from utils.graphs.line import LineGraph
from utils.graphs.bar import BarGraph
from utils.graphs.pie import PieGraph
from utils.currency_types import CurrencyType #MTD Received, MTD Expended, Month End Balance

#logging
import logging
import configparser

#other libraries
import time
from pathlib import Path

#emailing
from utils.email import send_email_to_user, send_completion_email, send_error_email


def config_setup():
    """Sets up the config file and creates the log file. It reads the config file and sets the input and output directories.

    Returns:
        _type_: returns the config settings
    """
    print("Setting up config file")
    try:
        config_dir = "logs"
        config_file = "config.ini"
        config = configparser.ConfigParser()
        config.read(config_file)
        return config
    except Exception as e:
        print(f"Error reading config file: {e}")
        logging.error(f"Error reading config file: {e}")

def logging_setup(config:configparser.ConfigParser):
    """Sets up the logging configuration. It creates a log file and sets the format of the log messages.
    Args:
        config (ConfigParser): The config file that contains the log file path and format
    """
    try:
        log_file = config['LOG_FILE']['info_log_path']
        format='%(asctime)s - %(levelname)s - %(filename)s:%(lineno)d - %(message)s'
        logging.basicConfig(filename=log_file, level=logging.INFO, format=format)
        return log_file
    except Exception as e:
        print(f"Error: {e} not found in config file.")
        logging.error(f"Error setting up logging: {e}")

def get_users_email(district_code:str, users_file:str)->list:

    df = pd.read_csv(users_file) #reads the users file
    filtered_df = df[df['District Code'] == district_code] #filters the dataframe for the specific district code
    if not filtered_df.empty:
        
        emails = filtered_df['Email'].tolist() #gets the email address for the district code
        #print(f"Emails found for district code {district_code}: {emails}")
        return emails

    else:
        print(f"No email found for district code: {district_code}")
        return [] #returns an empty list if no email is found for the district code

def get_all_districts(input_dir:str)->list:
    """Gets all the district directories in the input directory. It returns a list of all the district directories.

    Args:
        input_dir (str): directory path that contains all the district directories. Example: input/districts/

    Returns:
        list: of directory names. Example: ['NO', 'CN', 'EL']
    """
    district_dirs = os.listdir(parent_input_dir) #
    return district_dirs


def remove_input_files(district_code:str):
    """Removes the input files for the district code. It removes the input file and the formatted input file.
    This is used to remove the input files after the graphs have been created.

    Args:
        district_code (str): two letter district code. Example: NO, CN, etc. Used to identify the district's name and file paths
    """
    try:
        #Define all file paths in a single list
        base_path = os.path.join(parent_input_dir, district_code)

        files_to_delete = [
            # File 1: input_file
            os.path.join(base_path, config['INPUT_PATH']['input_file_name']),

            # File 2: formatted_input_file
            os.path.join(base_path, config['INPUT_PATH']['formatted_input_file']),

            # File 3: input_exp_file
            os.path.join(base_path, config['INPUT_PATH']['input_exp_pie_file']),

            # File 4: input_rev_file
            os.path.join(base_path, config['INPUT_PATH']['input_rev_pie_file']),
        ]

        #Loop through the list and delete files if they exist
        for file_path in files_to_delete:
            if os.path.exists(file_path):
                os.remove(file_path)
                print(f"Deleted existing file: {file_path}")

    except Exception as e:
        #Handle any errors during file removal
        print(f"Error removing input files: {e}")
        logging.error(f"Error removing input files for district {district_code}: {e}")
        send_error_email(e, config, district_code)



def process_exp_pie_chart(district_code:str):
    """Creates the pie charts for expenditures. It reads the input file and creates a new excel file for each account code from the input file.
    There is only one pie chart of expenditures in the output file for that account code. If there are multiple account codes in the input file, it will create muiltiple output files with one chart each.
    It finds the input file by building a file path using the district code and the input file name in the config.ini file.
    Args:
        district_code (str): two letter district code. Example: NO, CN, etc. Used to identify the district's name and file paths
    """

    print("\t\tstarting exp pie charts")

    #input file path for the pie chart input file. Example: input/NO/001-0000.csv
    try:
        input_exp_file = os.path.join(parent_input_dir, district_code, config['INPUT_PATH']['input_exp_pie_file'])
        
        #convert inputfile to a dataframe
        exp_df = pd.read_csv(input_exp_file)
        account_codes = exp_df['Cash Account'].unique().tolist() #list of all account codes in the pie chart input file

        try:
            for account_code in account_codes:

                #print(f"\t\tProcessing account code: {account_code} " + " for district: " + district_code)

                filtered_df = exp_df[exp_df['Cash Account'] == account_code] #filters the dataframe for just the specific account code
                output_file_name = account_code + " " + district_code + " " + config['OUTPUT_PATH']['graphs_file_name']  #output file name. Example: 001-0000 graphs.xlsx
                output_file_path = os.path.join(parent_output_dir, district_code, output_file_name) # Example output/NO/001-0000 graphs.xlsx

                #create file and load it's workbook if it exists. Else create a new workbook
                if os.path.exists(output_file_path):
                    #print(f" {output_file_path} file exists")
                    main_wb = openpyxl.load_workbook(output_file_path) 
                else:
                    #print(f" {output_file_path} file does not exist")
                    main_wb = Workbook()

                pie_title = "FYTD " + account_code + " Expenditures"
                new_exp_pie_ws = main_wb.create_sheet(title= "FYTD Expenditures Pie Chart")
                exp_pie_chart = PieGraph(district_code,input_exp_file, filtered_df, main_wb, account_code) #arg is the pie graph title           
                exp_pie_chart.create_chart(pie_title, new_exp_pie_ws, CurrencyType.Expended) #creates the pie graph using the PieGraph class. It will create a pie graph for each currency type        
                main_wb.save(output_file_path) #save the workbook with the pie charts
                excel_files_dict[account_code] = output_file_path #adds the file path to the dictionary for each account code

        except Exception as e:
            print(f"Error processing account codes for expenditure pie charts {e}")
            logging.error(f"Error processing account codes for expenditure pie charts: {e}")
            send_error_email(e, config, district_code)
        time.sleep(0.5) #sleep for 0.5 seconds to prevent too many requests at once

    except FileNotFoundError:
        print(f"Error: CSV file not found at '{input_exp_file}'.")
        logging.error(f"Error: CSV file not found at '{input_exp_file}'.")
        send_error_email(e, config, district_code)

def process_rev_pie_chart( district_code:str):
    print("\t\tProcessing rev pie charts")

    try:
        input_rev_file = os.path.join(parent_input_dir, district_code, config['INPUT_PATH']['input_rev_pie_file']) #input file path for the pie chart input file. Example: input/NO/001-0000.csv
        rev_df = pd.read_csv(input_rev_file)
        account_codes = rev_df['Cash Account'].unique().tolist() #list of all account codes in the pie chart input file

        try:
            for account_code in account_codes:
                #print(f"\tProcessing account code: {account_code}")

                filtered_df = rev_df[rev_df['Cash Account'] == account_code] #filters the dataframe for just the specific account code
                output_file_name = output_file_name = account_code + " " + district_code + " " + config['OUTPUT_PATH']['graphs_file_name']  #output file name. Example: 001-0000 graphs.xlsx
                output_file_path = os.path.join(parent_output_dir, district_code, output_file_name) # Example output/NO/001-0000 graphs.xlsx

                #create file and load it's workbook if it exists. Else create a new workbook
                if os.path.exists(output_file_path):
                    main_wb = openpyxl.load_workbook(output_file_path) 
                else:
                    main_wb = Workbook()

                pie_title = "FYTD " + account_code + " Revenue"
                new_rev_pie_ws = main_wb.create_sheet(title="FYTD Revenue Pie Chart")
                rev_pie_chart = PieGraph(district_code,input_rev_file, filtered_df, main_wb, account_code) #arg is the pie graph title
                rev_pie_chart.create_chart(pie_title, new_rev_pie_ws, CurrencyType.Revenue) #creates 
                main_wb.save(output_file_path) #save the workbook with the pie charts
                excel_files_dict[account_code] = output_file_path #adds the file path to the dictionary for each account code

        except Exception as e:
            print(f"Error processing account codes for revenue pie charts: {e}")
            logging.error(f"Error processing account codes for revenue pie charts: {e}")
            send_error_email(e, config, district_code)
        time.sleep(0.5) #sleep for 0.5 seconds to prevent too many requests at once
        
    except FileNotFoundError:
        print(f"Error: CSV file not found at '{input_rev_file}'.")
        logging.error(f"Error: CSV file not found at '{input_rev_file}'.")
        send_error_email(e, config, district_code)


def process_line_and_bar_graphs(district_code:str):
    """Creates the line and bar graphs for each account code in the input file. It reads the input file and creates a new excel file for each account code.
        It finds the districts input file by building a file path using the district code and the input file name in the config.ini file.
    Args:
        district_code (str): two letter district code. Example: NO, CN, etc. Used to identify the district's name and file paths
    """
    print("\tProcessing line AND bar graphs:")

    #dictionary containing the file paths for each district. File paths are listed in the config.ini file. If any errors occur make sure those file paths are correct
    #Note - this must remain in district for loop. Otherwise it will keep appending file paths rather than resetting them
    input_paths = {'cash_history': config['INPUT_PATH']['input_file_name'],
                   'formatted_input': config['INPUT_PATH']['formatted_input_file']
                    }
    try:
            
        input_paths['formatted_input'] = format_input_file(  
            os.path.join(parent_input_dir, district_code, input_paths['cash_history'])
            ,os.path.join(parent_input_dir, district_code, input_paths['formatted_input'])
            )
    
        #build a list of unique account codes. Need to filter dataframe for that specific account code
        formatted_file_df = pd.read_csv(input_paths['formatted_input'])
        #print(f"\tFormatted file dataframe shape: {formatted_file_df}")
        account_codes = formatted_file_df['Full Account Code'].unique().tolist()

        for account_code in account_codes:
                
                print(f"\t\tProcessing account code: {account_code}")
                main_wb = Workbook() #create a new workbook. This must be in loop so it creates a new excel file for each account code
                output_file_name_only = account_code + " " + district_code + " " + config['OUTPUT_PATH']['graphs_file_name']
                output_file_path = os.path.join(parent_output_dir, district_code, output_file_name_only) # Example output/NO/001-0000 graphs.xlsx

                #filters the dataframe for just the specific account code
                filtered_df = formatted_file_df[formatted_file_df['Full Account Code'] == account_code] 

                #creates the line graph for month end balances
                process_line_graph(district_code, input_paths['formatted_input'], main_wb, filtered_df , account_code)

                #creates the bar graphs for each currency type
                process_bar_graphs(district_code, input_paths['formatted_input'], main_wb, filtered_df, account_code) 
            
                try:                
                    #print("\t\tsaving graphs file to this file path: ", output_file_path)
                    main_wb.save(output_file_path) 
                    excel_files_dict[account_code] = output_file_path #adds the file path to the dictionary for each account code

                except Exception as e:
                    logging.error(f"Error saving output file: {e}")
                    print(f"Error saving output file: {e}")   
                    send_error_email(e, config, district_code)

    except Exception as e:
        print(f"Error input file non found for district: {district_code}: {e}")
        logging.error(f"Error formatting input file for district {district_code}: {e}")
        send_error_email(e, config, district_code)




def process_bar_graphs(district_code:str, input_file:str, current_workbook:Workbook, input_df:DataFrame, account_code:str):
    """Creates a bar graph for each currency type. The currency types are MTD Received, MTD Expended, and Month End Balance.
    It will create 3 graphs total for each account code. It will create a new worksheet for each graph.

    Args:
        district_code (str): two letter district code. Example: NO, CN, etc. Used to identify the district's name and file paths
        input_file (str): path to the input file. Example: input/NO/001-0000.csv
        current_workbook (Workbook): the current workbook that the graphs will be created in
        input_df (DataFrame): filtered dataframe for the specific account code
        account_code (str): Fund-scc code for the cash account. Example: 001-0000
    """
    print(f"\t\tProcessing bar graphs for district: {district_code} and account code: {account_code}")
    try:

        #It creates a graph for each currency type. The currency types are MTD Received, MTD Expended, and Month End Balance
        for currencyType in CurrencyType:
           
            new_bar_ws = current_workbook.create_sheet("Bar Graph - " + currencyType.value) #create new worksheet in current excel file
            new_bar_chart = BarGraph(district_code, input_file, input_df, current_workbook, account_code) #arg is the bar graphs title
            bar_title = account_code + " " + currencyType.value #sets the title of the bar graph to be the account code and the currency type
            new_bar_chart.create_chart(bar_title, new_bar_ws, currencyType) #creates the bar graph using the BarGraph class. It will create a bar graph for each currency type
            time.sleep(0.5) #sleep for 0.5 seconds to prevent too many requests at once

    except FileNotFoundError:
        print(f"Error: CSV file not found at '{input_file}'.")


def process_line_graph(district_code:str, input_file:str, current_workbook:Workbook, input_df:DataFrame, account_code:str):
    """Creates a line graph for month end expenditures, revenue, and ending balances. It creates the graph on the initial worksheet of the workbook.

    Args:
        district_code (str): two letter district code. Example: NO, CN, etc. Used to identify the district's name and file paths
        input_file (str): path to the input file. Example: input/NO/001-0000.csv
        current_workbook (Workbook): the current workbook that the graphs will be created in
        input_df (DataFrame): filtered dataframe for the specific account code
        account_code (str): Fund-scc code for the cash account. Example: 001-0000
    """

    print(f"\t\t\tProcessing line graph from file: {input_file} for district: {district_code} and account code: {account_code}")
    try:

        #new_line_ws = workbook.create_sheet(title="Line Graph - Month End Balances")
        new_line_ws = current_workbook.active #active is the first worksheet in the workbook. It will be used for the line graph
        new_line_ws.title = " Month End Balances" #sets the title of the worksheet to be the same as the graph title
        line_title = account_code + " Month End Balances"
        #new_line_chart = LineGraph(district_code, input_file, input_df, current_workbook, account_code) 
        new_line_chart = LineGraph(district_code, input_file, input_df, current_workbook, account_code) #arg is the line graph title
        new_line_chart.create_chart(line_title, new_line_ws, currency_type=None) #currency type is not needed for this line graph as it will plot all currency types anyways. Arguement should be set to none when called.
    
    except FileNotFoundError:
        print(f"Error: CSV file not found at '{input_file}'.")
    
    time.sleep(0.5) #sleep for 0.5 seconds to prevent too many requests at once


#main function
if __name__ == "__main__":

    print("Starting Excel Graph Program")
    #print(openpyxl.__version__) #print the version of openpyxl

    #setup logging and config file
    config = config_setup()
    log_file = logging_setup(config)
    logging.info(f"Log file created at {log_file}") #logs when file was create and when program was last run
    admin_email = str(config['EMAIL_SETTINGS']['admin_email_address'])
    try:    
        parent_input_dir = config['INPUT_PATH']['input_dir'] #input/districts/ directory
        parent_output_dir = config['OUTPUT_PATH']['output_dir'] #output/districts/ directory         
        district_dirs = get_all_districts(parent_input_dir) #list of all districts in the input/districts directory

    except Exception as e:
        logging.error(f"Error getting district directories: {e}")
    
    #iterate through each district directory and get the input file names
    for district_code in district_dirs:
        print(f"Processing district: {district_code}")
        excel_files_dict = {} #dictionary to hold the excel files for each district. Key is the account code and value is the file path
        time.sleep(1) #sleep for 1 second to prevent too many requests at once
        
        process_line_and_bar_graphs(district_code)
        process_exp_pie_chart(district_code) #creates the pie charts for expenditures
        process_rev_pie_chart(district_code) #creates the pie charts for revenue
        remove_input_files(district_code) #removes the input files for the district code
        
        
        for user in get_users_email(district_code, config['USERS']['users_file']):
            #print(f"Sending email to user: {user}")
            send_email_to_user(district_code, excel_files_dict, user, config)
        
    
    print("\nFinished Graphs For All Districts")
    send_completion_email(config) #sends an email to me
   