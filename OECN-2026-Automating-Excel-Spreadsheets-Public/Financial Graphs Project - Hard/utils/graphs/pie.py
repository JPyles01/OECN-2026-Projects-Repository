import pandas as pd
import traceback
from pandas import DataFrame
from typing import List
from openpyxl import Workbook, worksheet
from openpyxl.chart import PieChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
import logging
from logging import exception
from ..data_processing import format_to_currency, auto_adjust_column_width

from ..currency_types import CurrencyType
from ..base_chart import BaseChart

#allows logging to be used in this file
logger = logging.getLogger(__name__)

class PieGraph(BaseChart):
    #dimensions for the pie charts. These can be adjusted as needed. The default is set to 25 width and 13 height which seems to be a good size for the charts.
    width = 25
    height = 13


    #these are the default expenditures object levels. Headers. Originally was placed outside of function. 
    EXP_CATEGORIES_DICT = {100:"Salaries and Wages", 200:"Benefits", 400:"Purchased Services", 500:"Supplies", 600:"Capital Outlay", 800:"Other Objects", 900:"Other Uses of Funds"} #This is a dictionary that contains the object levels and their descriptions. Used for testing purposes only.

    #The categories for the revenue pie chart 
    REVENUE_CATEGORIES = ["State", "Federal", "Real Estate", "Personal Property", "School District Income Tax","All Other General Fund Revenue"] 
    
    REV_CATEGORIES_DICT = {3000:"State", 4000:"Federal", 1111:"Real Estate", 1122:"Personal Property", 1130:"School District Income Tax", "Other":"All Other Fund Revenue"} #This is a dictionary that contains the revenue categories and their descriptions. Used for testing purposes only.
    
    def __init__(self, district_code:str, input_file:str, input_df:DataFrame, workbook:Workbook, account_code:str):
        super().__init__(district_code, input_file, input_df, workbook, account_code)


    def create_chart(self, chart_title:str, current_ws:worksheet, currency_type:CurrencyType):
        """Creates a pie chart based on the currency type provided.

        Args:
            chart_title (str): Tile of the chart
            current_ws (worksheet): Worksheet in the current excel file where the chart will be created in.
            currency_type (CurrencyType): Currency type for the chart. It can be Expended, Revenue, or Balances.
        """


        if currency_type == CurrencyType.Expended:
            self.create_exp_pie_chart(chart_title, current_ws, self.input_file)

        if currency_type == CurrencyType.Revenue:
            self.create_revenue_pie_chart(chart_title, current_ws, self.input_file)
        
        if currency_type == CurrencyType.Balances:
            print("Balances currency type cannot be used for Pie Charts. Please select expenditures or revenue.")

    def create_exp_pie_chart(self, chart_title:str, current_ws: worksheet, csv_file: str):
        """Creates a pie chart based on expenditures data from a CSV file.
        This function reads a CSV file containing expenditure data, aggregates it by object level, and creates a pie chart

        Args:
            current_ws (worksheet): worksheet in the current excel file where the chart will be created in.
            csv_file (str): file path to the CSV file containing expenditure data.

        Raises:
            ValueError: If the required columns are missing in the CSV file.
        """
        ws = current_ws

        # Load the CSV file into a DataFrame
        pie_df = pd.read_csv(csv_file)

        # Check for required columns
        required_columns = ["Object Level", "FYTD Expended"]
        for column in required_columns:
            if column not in pie_df.columns:
                raise ValueError(f"Missing required column '{column}' in the CSV file.")

        # Subtotals by each unique object level and places those totals into a new DataFrame
        aggregated_data = pie_df.groupby("Object Level")["FYTD Expended"].sum().reset_index()

        ws.cell(row=1, column=1, value="Object Level")
        ws.cell(row=1, column=2, value="Description")
        ws.cell(row=1, column=3, value="FYTD Expended")  

        
        for r in range(aggregated_data.shape[0]):
            object_level = aggregated_data["Object Level"][r]
            ws.cell(row=r + 2, column=1, value=aggregated_data["Object Level"][r])
            #ws.cell(row=r + 2, column=2, value=self.EXP_CATEGORIES[r])
            ws.cell(row=r + 2, column=2, value=self.EXP_CATEGORIES_DICT[object_level]) #testing the dictionary option
            ws.cell(row=r + 2, column=3, value=aggregated_data["FYTD Expended"][r])

        format_to_currency(ws, min_column=2)
        auto_adjust_column_width(ws)

        # Create a pie chart
        pie = PieChart()
        pie.title = chart_title

        pie.width = self.width
        pie.height = self.height
        pie.style = 10

        # Pie chart's positioning and size
        pie.layout = Layout(ManualLayout(x=0, y=0, h=0.88, w=0.75))  # Shrinks the graph slightly so the legend does not overlap the chart

        # Title text positioning
        title_layout = ManualLayout(xMode="edge", yMode="edge", x=0.75, y=1.0)
        pie.title.layout = Layout(manualLayout=title_layout)

        # labels = Reference(ws, min_col=2, min_row=2, max_row=len(self.EXP_CATEGORIES) + 1)  # Object level descriptions
        # data = Reference(ws, min_col=3, min_row=1, max_row=len(self.EXP_CATEGORIES) + 1)        
        labels = Reference(ws, min_col=2, min_row=2, max_row = aggregated_data.shape[0] + 1)  # Object level descriptions
        data = Reference(ws, min_col=3, min_row=1, max_row = aggregated_data.shape[0] + 1)

        pie.add_data(data, from_rows=False, titles_from_data=True)
        pie.set_categories(labels)

        # Add labels to chart
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.number_format = '0.0%'
        pie.dataLabels.showCatName = True
        pie.dataLabels.showLegendKey = True
        pie.dataLabels.showSerName = False
        pie.dataLabels.showVal = True
        pie.dataLabels.separator = "\n"
        pie.dataLabels.dLblPos = 'bestFit'  # Positioning of the data labels, can be 'bestFit', 'outEnd', 'inEnd', etc.
        pie.dataLabels.showLeaderLines = True  # Show leader lines for better visibility

        # Add the chart to the worksheet
        ws.add_chart(pie, "E2")
        print("\t\t\tFinished creating expenditure pie chart for district: " + self.district_code + " and account code: " + self.account_code)
        logging.info("  Finished creating expenditure pie chart for district: " + self.district_code + " and account code: " + self.account_code)

    def create_revenue_pie_chart(self,chart_title:str, current_ws: worksheet, csv_file: str):
        print("\t\tStarting to create revenue...")

        ws = current_ws

        # Load the CSV file into a DataFrame
        rev_df = pd.read_csv(csv_file)

        # Check for required columns
        required_columns = ["Receipt One Digit Level", "Receipt", "FYTD Received"]
        for column in required_columns:
            if column not in rev_df.columns:
                raise ValueError(f"Missing required column '{column}' in the CSV file.")

        # Define helper functions
        def total_by_receipt_level(df: DataFrame, rc_level: int) -> float:
            filtered_df = df[df['Receipt One Digit Level'] == rc_level]
            total = filtered_df['FYTD Received'].sum()
            return total

        def total_by_specific_receipt_code(df: DataFrame, rc_code: int) -> float:
            filtered_df = df[df['Receipt'] == rc_code]
            total = filtered_df['FYTD Received'].sum()
            return total

        def calculate_other_revenue(df: DataFrame, exclude_codes: List[int]):
            # Filter the DataFrame to exclude the specified codes
            filtered_df = df[~df['Receipt'].isin(exclude_codes)]  # The ~ inverts the boolean so it becomes True for the codes that are NOT in the exclude_codes
            revenue = filtered_df['FYTD Received'].sum()
            return revenue

        def get_list_of_receipts_by_level(df: DataFrame, rc_level: int) -> List[int]:
            receipts = []
            filtered_df = df.loc[df['Receipt One Digit Level'] == rc_level, 'Receipt']
            receipts = filtered_df.tolist()
            return receipts

        # Continue with the rest of the method logic...

        ws.cell(row=1, column=1, value = "Revenue Categories")
        ws.cell(row=1, column=2, value = "FYTD Received")


        exclude_rc = [1111,1122,1130] #list of receipt codes that will NOT be used to calculate Other revenue

        #state receipts are 3000 level codes
        state_total = total_by_receipt_level(rev_df, 3000)
        #state_desc = self.REVENUE_CATEGORIES[0] #State
        state_desc = self.REV_CATEGORIES_DICT[3000] #State description from the dictionary
        ws.cell(row=2, column=1, value = state_desc)
        ws.cell(row=2, column=2, value = state_total)

        #federal receipts are 4000 level codes
        fed_total = total_by_receipt_level(rev_df, 4000)
        #fed_desc = self.REVENUE_CATEGORIES[1]
        fed_desc = self.REV_CATEGORIES_DICT[4000] #Federal description from the dictionary
        ws.cell(row=3, column=1, value = fed_desc)
        ws.cell(row=3, column=2, value = fed_total)
        

        #real estate receipt code is specifically 1111
        real_estate_total = total_by_specific_receipt_code(rev_df, 1111)
        #real_estate_desc = self.REVENUE_CATEGORIES[2]
        real_estate_desc = self.REV_CATEGORIES_DICT[1111] #Real Estate description from the dictionary
        ws.cell(row=4, column=1, value = real_estate_desc)
        ws.cell(row=4, column=2, value = real_estate_total)
        
        #personal property receipt code is specifically 1122
        personal_prop_total = total_by_specific_receipt_code(rev_df, 1122)
        #personal_prop_desc = self.REVENUE_CATEGORIES[3]
        personal_prop_desc = self.REV_CATEGORIES_DICT[1122] #Personal Property description from the dictionary

        ws.cell(row=5, column=1, value = personal_prop_desc)
        ws.cell(row=5, column=2, value = personal_prop_total)

        #OSDI revenue is specifcally 1130. Not all districts have OSDI income tax. For example, an ESC may not have any
        osdi_total = total_by_specific_receipt_code(rev_df, 1130)
        #osdi_desc = self.REVENUE_CATEGORIES[4]
        osdi_desc = self.REV_CATEGORIES_DICT[1130] #OSDI description from the dictionary
        ws.cell(row=6, column=1, value = osdi_desc)
        ws.cell(row=6, column=2, value = osdi_total)

        #below builds a list of receipt codes to not include in the other revenue category.
        #the receipts not to use are all 3000 and 4000 codes as well as specific codes 1111,1122, and 1130
        state_recs = get_list_of_receipts_by_level(rev_df, 3000) 
        federal_recs = get_list_of_receipts_by_level(rev_df, 4000)

        exclude_rc += state_recs
        exclude_rc += federal_recs

        other_revenue_total = calculate_other_revenue(rev_df, exclude_rc)
        #other_rev_desc = self.REVENUE_CATEGORIES[5]
        other_rev_desc = "ALL OTHER FUND REVENUE" #This is the description for the other revenue category
        ws.cell(row=7, column=1, value = other_rev_desc)
        ws.cell(row=7, column=2, value = other_revenue_total)

        format_to_currency(ws, min_column=2)
        auto_adjust_column_width(ws)
        
        pie = PieChart()
        pie.title = chart_title

        #pie chart's positioning and size
        pie.layout = Layout(ManualLayout(x = 0, y = 0, h = 0.88, w = 0.75)) #This shrinks the graph slightly so the legend does not overlap the actualy chart

        #title text positioning
        title_layout = ManualLayout(xMode="edge",yMode="edge",x=0.75,y=1.0)
        pie.title.layout = Layout(manualLayout=title_layout)

        pie.width = self.width
        pie.height = self.height
        pie.style = 10

        labels = Reference(ws, min_col=1, min_row=2, max_row=len(self.REV_CATEGORIES_DICT) + 1)
        data = Reference(ws, min_col=2,min_row=1, max_row=len(self.REV_CATEGORIES_DICT) + 1)

        pie.add_data(data, from_rows=False, titles_from_data=True)
        pie.set_categories(labels)
       
         #Add Labels to chart
        pie.dataLabels = DataLabelList()
        #pie.dataLabels.show_val = True
        pie.dataLabels.showPercent = True
        pie.dataLabels.numFmt = '0.0%'
        pie.dataLabels.showCatName = True
        pie.dataLabels.showSerName = False
        pie.dataLabels.showLegendKey = True
        pie.dataLabels.showVal = True
        pie.dataLabels.separator = "\n"
        pie.dataLabels.dLblPos = 'bestFit'
        pie.dataLabels.showLeaderLines = True  # Show leader lines for better visibility

        ws.add_chart(pie, "E2")
        print("\t\t\tFinished creating revenue pie chart for district: " + self.district_code + " and account code: " + self.account_code)
        logging.info("  Finished creating revenue pie chart for district: " + self.district_code + " and account code: " + self.account_code)





