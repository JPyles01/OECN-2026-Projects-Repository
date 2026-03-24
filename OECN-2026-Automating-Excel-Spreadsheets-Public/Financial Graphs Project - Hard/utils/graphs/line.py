import pandas as pd
from pandas import DataFrame
from openpyxl import Workbook, worksheet, load_workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.styles import Color
from openpyxl.drawing.colors import ColorChoice
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.marker import Marker
from openpyxl.utils.dataframe import dataframe_to_rows

import logging
from logging import exception
from ..data_processing import format_to_currency, auto_adjust_column_width, create_pivot_dataframe, sort_fiscal_years_by_months, create_monthly_dataframe
from ..currency_types import CurrencyType


from ..base_chart import BaseChart

#allows logging to be used in this file
logger = logging.getLogger(__name__)

class LineGraph(BaseChart):

    width = 28
    height = 12
    marker_symbols = ["circle", "square", "x"]
    prior_years_months = ["J", "A", "S", "O", "N", "D", "J", "F", "M", "A", "M", "J", "J", "A", "S", "O", "N", "D", "J", "F", "M", "A", "M", "J", "J", "A", "S", "O", "N", "D", "J", "F", "M", "A", "M", "J"] # 3 prior fiscal year months

    def __init__(self, district_code:str, input_file:str, input_df:DataFrame, workbook:Workbook, account_code:str):
        """Initializes the LineGraph class with district code, input file, input dataframe, workbook"""

        super().__init__(district_code, input_file, input_df, workbook, account_code ) #I could probably remove either the input file or input_df since they contain the same information
       

    def create_chart(self, chart_title:str, current_ws:worksheet, currency_type:CurrencyType):
        """Create a line chart containing expenditure, revenue, and ending balance data for the current and prior 3 fiscal years. There will be 3 different lines in this chart reflecting each type of data.
        This will create the necessary dataframes and plot each type of data.
        It will write the data onto the excel spreadsheet.
        It will then use that data to create a line graph

        Args:
            current_ws (worksheet): The worksheet in excel the graph will be created on and have data written to
            currency_type (CurrencyType): This is not necessary for this line graph as it will plot all currency types anyways. Arguement should be set to none when called.
            input_file (str): The csv file that contains the data that will be used.
        """

        chart = LineChart()
        ws = current_ws

        expended_pivot_df = create_pivot_dataframe(self.input_df, CurrencyType.Expended)
        expended_sorted_df = sort_fiscal_years_by_months(expended_pivot_df)

        #MTD Revenue Data
        revenue_pivot_df = create_pivot_dataframe(self.input_df, CurrencyType.Revenue)
        revenue_sorted_df = sort_fiscal_years_by_months(revenue_pivot_df)

        #MTD Ending Balances Data
        month_end_pivot_df = create_pivot_dataframe(self.input_df,CurrencyType.Balances)
        month_end_sorted_df = sort_fiscal_years_by_months(month_end_pivot_df)

        #Builds lists containing just the data that will be written to the spreadsheet
        month_names = expended_sorted_df.columns.tolist() #Builds a list of all month names in fiscal year order
        expended_data = expended_sorted_df.values.flatten().tolist() #builds a list of all expended amounts
        revenue_data = revenue_sorted_df.values.flatten().tolist()
        month_end_data = month_end_sorted_df.values.flatten().tolist()


        # Create 48-month x-axis labels
        years = expended_pivot_df.index.tolist() #builds a list of fiscal years
        all_months = [f"{month} {year}" for year in years for month in month_names]

        ##This commeted out code is if the data should be written in a horizontal layout where each row is a currency type
        ##This layout there would be 49 columns total. Column 1 are the titles and each column is a month
        # #This method creates 2 rows with 48 columns of data - horizontal layout
        # for col_num, month in enumerate(all_months, start=2):
        #     ws.cell(row=1, column=col_num, value = month)
        
        # for col_num, value in enumerate(expended_data, start=2):
        #     ws.cell(row=2, column=col_num, value=value)

        # for col_num, value in enumerate(revenue_data, start=2): #start at column 2
        #     ws.cell(row=3, column=col_num, value=value)

        # for col_num, value in enumerate(month_end_data, start=2): #start at column 2
        #     ws.cell(row=4, column=col_num, value=value)

        #This is for vertical layout. If Horizontal layout the rows and columns will need to be changed
        ws.cell(row = 1, column = 1, value = "Months")
        ws.cell(row = 1, column = 2, value = "MTD Expenditures")
        ws.cell(row = 1, column = 3, value = "MTD Received")
        ws.cell(row = 1, column = 4, value = "MTD Ending Balances")
   
        for month, expended, revenue, balance in zip(all_months, expended_data, revenue_data, month_end_data):
            ws.append([month, expended, revenue, balance])

        #horizontal orientation
        # months_only = Reference(ws, min_row = 1, max_row=1, min_col=2, max_col=ws.max_column)
        # expended_only = Reference(ws, min_col=1, max_col=49, min_row=2, max_row=2)
        # revenue_only = Reference(ws,min_col=1, max_col=49, min_row=3, max_row=3)
        # month_balances_only = Reference(ws,min_col=1, max_col=49, min_row=4, max_row=4)


        #vertical Orientation
        months_only = Reference(ws, min_row = 2, max_row=ws.max_row, min_col=1, max_col=1)
        expended_only = Reference(ws, min_row = 1, max_row=ws.max_row, min_col=2, max_col=2)
        revenue_only = Reference(ws, min_row = 1, max_row=ws.max_row, min_col=3, max_col=3)
        balances_only = Reference(ws, min_row = 1, max_row=ws.max_row, min_col=4, max_col=4)


        #if doing horizontal orientation, then from_rows must be set to True
        chart.add_data(expended_only, from_rows=False, titles_from_data = True)
        chart.add_data(revenue_only, from_rows=False, titles_from_data = True)
        chart.add_data(balances_only, from_rows=False, titles_from_data = True)

        #Sets x axis labels for line chart
        chart.set_categories(months_only)

        auto_adjust_column_width(ws) #adjusts column widths for large numbers
        format_to_currency(ws,min_column=2) #changes numbers into currency format
        
        chart.style = 1
        chart.width = self.width
        chart.height = self.height

        #set chart title and legend position
        chart.title = chart_title          
        chart.legend.position = 'r'
        chart.legend.overlay = True

        # Add tick marks to the x-axis
        chart.x_axis.majorTickMark = "out"  # "in", "out", "cross", "none"
        chart.x_axis.minorTickMark = "out" # "in", "out", "cross", "none"
        chart.y_axis.title = "Amounts"
        chart.x_axis.delete = False #enables the x axis labels - Months
        chart.y_axis.delete = False #enables the y axis labels - Dollar amount

        #changes expenditures to red line
        chart.series[0].graphicalProperties.line.solidFill = ColorChoice(prstClr='indianRed')
        marker1 = Marker(symbol="circle", spPr=GraphicalProperties(solidFill=ColorChoice(prstClr='indianRed')))  
        chart.series[0].marker = marker1 #set marker to circle

        #changes revenue to green line
        chart.series[1].graphicalProperties.line.solidFill = ColorChoice(prstClr = 'seaGreen')
        marker2 = Marker(symbol="square", spPr=GraphicalProperties(solidFill=ColorChoice(prstClr='seaGreen')))
        chart.series[1].marker = marker2 #set marker to square

        
        #changes ending balances to blue
        chart.series[2].graphicalProperties.line.solidFill = ColorChoice(prstClr = 'medBlue')
        marker3 = Marker(symbol="x", spPr=GraphicalProperties(solidFill=ColorChoice(prstClr='medBlue')))
        chart.series[2].marker = marker3 #set marker to x

        ws.add_chart(chart, "F2")  # Position the chart in cell D2
        print("\t\t\tFinished creating line chart for district: " + self.district_code + " and account code: " + self.account_code)
        logging.info("  Finished creating line chart for district: " + self.district_code + " and account code: " + self.account_code)



    #Example of how to create a line chart
    def example_line_chart(self, output_file:str):
        print("Example Chart")
        # Sample data
        data = [
            ["Month", "2023"],
            ["Jan", 10],
            ["Feb", 15],
            ["Mar", 12],
            ["Apr", 18],
            ["May", 20],
        ]

        wb = Workbook()
        ws = wb.active

        for row in data:
            ws.append(row)

        data_values = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=len(data))
        categories = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=len(data))

        chart = LineChart()
        chart.title = "Example Line Chart"
        chart.style = 1
        chart.add_data(data_values)
        chart.set_categories(categories)

        ws.add_chart(chart, "A1")  # Place chart in Cell A1

        wb.save(output_file)

        print("chart_styles.xlsx created. Open it to see the styles.")

